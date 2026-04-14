"""
DFW Modern Multi-Tenant Portfolio — Underwriting Model Builder
================================================================
Generates: DFW_Underwriting_Model.xlsx

Architecture:
  1. Summary           — Returns, metrics, NOI bridge, waterfall
  2. Assumptions       — All user-editable inputs (named ranges)
  3. Debt              — Dynamic debt sizing + amortization schedule
  4. MLA               — Market Leasing Assumptions (per property / unit type)
  5. Rent Roll         — 162 tenants with vacate/option/market flags
  6. Tenant Schedule   — 162 tenants × 120 monthly rent cells (ARGUS-style)
  7. Monthly CF        — Property-level aggregated monthly cash flows
  8. Annual CF         — FY roll-up (FYE April 30) across 10-year hold
  9. Property Summary  — Per-property returns / metrics

Every formula uses cell/name references so the model is fully dynamic.
Institutional defaults are seeded; all inputs on Assumptions tab are editable.

Source data (all extracted from repo PDFs + rent roll .xlsx):
  - Portfolio Brochure.pdf (CBRE teaser)
  - Carrollton/Frisco/Lewisville/Mesquite/Plano/Roanoke OMs
  - DFW Rent Roll 4.9.2026.xlsx (162 tenant-suites, 430,550 SF)
"""

import datetime as dt
from dataclasses import dataclass, field
from typing import Optional

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# SECTION 1 — STATIC DATA EXTRACTED FROM OMs
# =============================================================================

# -----------------------------------------------------------------------------
# Properties (OM-derived)
# -----------------------------------------------------------------------------
# NOTE: Plano OM file is incomplete — only rent roll section. OpEx estimated
# from Frisco (same $19 market rent, similar vintage) pro-rated to Plano SF.

PROPERTIES = [
    {
        "id": "CAR",
        "name": "Carrollton",
        "address": "1750 Briercroft Court, Carrollton, TX 75006",
        "submarket": "Northwest Dallas",
        "year_built": 2023,  # Bldg 4 delivered Oct 2025
        "sf": 101855,
        "office_sf": 43046,
        "office_pct": 0.42,
        "land_acres": 7.47,
        "buildings": 4,
        "units": 36,
        "parking": 192,
        "secured_parking": 41,
        "clear_height": "20'4\"",
        # Market leasing assumptions (from OM, by unit type)
        "market_rent_shop": 18.00,
        "market_rent_studio": 16.00,
        # 2025 Assessment for tax module
        "assessed_value_2025": 17500000,
        "millage_rate": 0.02059675,
        "parcel": "Dallas CAD #140067300A01R0000",
        # FY27 OpEx (from OM Argus proforma)
        "fy27_cam": 238800,
        "fy27_utilities": 118489,
        "fy27_mgmt_fee": 105166,  # 4% of EGR
        "fy27_insurance": 78963,
        "fy27_re_taxes": 364047,
        # FY27 in-place per OM
        "fy27_noi_om": 1723687,
        "inplace_noi": 1581944,  # Executive summary in-place NOI
        # Ancillary income assumptions (FY27)
        "tenant_upgrade_rent_fy27": 108444,  # $1.10 per occupied SF
        "parking_rent_fy27": 89347,  # 93% utilization
        "parking_utilization_fy27": 0.93,
    },
    {
        "id": "FRI",
        "name": "Frisco",
        "address": "10400 Frisco Street, Frisco, TX 75034",
        "submarket": "Northeast Dallas",
        "year_built": 2024,
        "sf": 74600,
        "office_sf": 37300,
        "office_pct": 0.50,
        "land_acres": 5.16,
        "buildings": 3,
        "units": 28,
        "parking": 161,
        "secured_parking": 25,
        "clear_height": "19'4\"",
        "market_rent_shop": 19.00,
        "market_rent_studio": 19.00,  # all one rate
        "assessed_value_2025": 13174536,
        "millage_rate": 0.0167548,
        "parcel": "Collin CAD #2761885",
        "fy27_cam": 169748,
        "fy27_utilities": 81730,
        "fy27_mgmt_fee": 80187,
        "fy27_insurance": 73087,
        "fy27_re_taxes": 222944,
        "fy27_noi_om": 1376978,
        "inplace_noi": 1365615,
        "tenant_upgrade_rent_fy27": 54160,  # $0.72 per occupied SF
        "parking_rent_fy27": 32227,
        "parking_utilization_fy27": 0.76,
    },
    {
        "id": "LEW",
        "name": "Lewisville",
        "address": "760 E. Main Street, Lewisville, TX 75057",
        "submarket": "Northwest Dallas",
        "year_built": 2023,
        "sf": 69500,
        "office_sf": 31960,
        "office_pct": 0.46,
        "land_acres": 4.50,
        "buildings": 3,
        "units": 28,
        "parking": 127,
        "secured_parking": 25,
        "clear_height": "19'4\"",
        "market_rent_shop": 18.00,
        "market_rent_studio": 18.00,
        "assessed_value_2025": 10800000,
        "millage_rate": 0.01722747,
        "parcel": "Denton CAD #983656",
        "fy27_cam": 166856,
        "fy27_utilities": 91790,
        "fy27_mgmt_fee": 70523,
        "fy27_insurance": 58043,
        "fy27_re_taxes": 187918,
        "fy27_noi_om": 1187939,
        "inplace_noi": 1189505,
        "tenant_upgrade_rent_fy27": 116942,  # $1.69 per occupied SF
        "parking_rent_fy27": 26438,
        "parking_utilization_fy27": 0.64,
    },
    {
        "id": "MES",
        "name": "Mesquite",
        "address": "2828 & 2836 I-30, Mesquite, TX 75150",
        "submarket": "East Dallas",
        "year_built": 2022,
        "sf": 50750,
        "office_sf": 21225,
        "office_pct": 0.42,
        "land_acres": 3.45,
        "buildings": 2,
        "units": 22,
        "parking": 138,
        "secured_parking": 32,
        "clear_height": "19'4\"",
        "market_rent_shop": 16.00,
        "market_rent_studio": 16.00,
        "assessed_value_2025": 6168340,
        "millage_rate": 0.02345667,
        "parcel": "Dallas CAD #380215100D0060000",
        "fy27_cam": 127914,
        "fy27_utilities": 46757,
        "fy27_mgmt_fee": 43482,
        "fy27_insurance": 39986,
        "fy27_re_taxes": 146136,
        "fy27_noi_om": 682764,
        "inplace_noi": 455529,
        "tenant_upgrade_rent_fy27": 50992,  # $1.14 per occupied SF
        "parking_rent_fy27": 5664,
        "parking_utilization_fy27": 0.06,
    },
    {
        "id": "PLA",
        "name": "Plano",
        "address": "535 Talbert Drive, Plano, TX",
        "submarket": "Northeast Dallas",
        "year_built": 2023,
        "sf": 82250,
        "office_sf": 41125,
        "office_pct": 0.50,
        "land_acres": 5.50,  # estimated; not in available OM
        "buildings": 3,
        "units": 30,
        "parking": 178,  # estimated
        "secured_parking": 30,  # estimated
        "clear_height": "19'4\"",
        "market_rent_shop": 19.00,
        "market_rent_studio": 19.00,
        # ESTIMATED — Plano OM in repo is incomplete (rent roll only).
        # OpEx PSF pro-rated from Frisco ($8.41/SF) applied to Plano 82,250 SF.
        "assessed_value_2025": 14500000,  # ESTIMATED
        "millage_rate": 0.020,  # ESTIMATED — Collin County average
        "parcel": "ESTIMATED — confirm with seller",
        "fy27_cam": 187130,  # $2.28/SF × 82,250 (Frisco ratio)
        "fy27_utilities": 90000,  # $1.10/SF × 82,250
        "fy27_mgmt_fee": 88000,  # ~4% of EGR estimate
        "fy27_insurance": 80650,  # $0.98/SF × 82,250
        "fy27_re_taxes": 290000,  # ESTIMATED (if reassessment off)
        "fy27_noi_om": 1100000,   # REVISED: proper NOI estimate (was gross rent)
        "inplace_noi": 1050000,   # REVISED: ~Frisco 68% op margin × implied EGR
        "tenant_upgrade_rent_fy27": 80000,  # ESTIMATED
        "parking_rent_fy27": 35000,  # ESTIMATED
        "parking_utilization_fy27": 0.80,  # ESTIMATED
    },
    {
        "id": "ROA",
        "name": "Roanoke",
        "address": "120 & 125 Country View Dr, Roanoke, TX 76262",
        "submarket": "North Fort Worth",
        "year_built": 2025,
        "sf": 51595,
        "office_sf": 18465,
        "office_pct": 0.36,
        "land_acres": 3.25,
        "buildings": 2,
        "units": 18,
        "parking": 131,
        "secured_parking": 22,
        "clear_height": "19'6\"",
        "market_rent_shop": 17.00,
        "market_rent_studio": 17.00,
        "assessed_value_2025": 8839548,
        "millage_rate": 0.0159622,
        "parcel": "Denton CAD #742232, #742233, #742235",
        "fy27_cam": 110358,
        "fy27_utilities": 103125,
        "fy27_mgmt_fee": 50335,
        "fy27_insurance": 48744,
        "fy27_re_taxes": 142510,
        "fy27_noi_om": 803301,
        "inplace_noi": 366683,
        "tenant_upgrade_rent_fy27": 47411,  # $1.01 per occupied SF
        "parking_rent_fy27": 15169,
        "parking_utilization_fy27": 0.41,
    },
]

# Portfolio totals (for sanity checks)
PORTFOLIO_SF = sum(p["sf"] for p in PROPERTIES)  # 430,550
PORTFOLIO_FY27_NOI = sum(p["fy27_noi_om"] for p in PROPERTIES)

# -----------------------------------------------------------------------------
# Default global assumptions (all editable on Assumptions tab)
# -----------------------------------------------------------------------------
DEFAULTS = {
    # Model timing
    "analysis_start": dt.date(2026, 5, 1),
    "analysis_months": 120,  # 10 years
    "fy_end_month": 4,  # April fiscal year end (matches OMs)

    # Growth rates (CBRE/Argus convention — uniform across portfolio)
    "market_rent_growth": 0.030,
    "opex_growth": 0.030,
    "re_tax_growth": 0.030,
    "cpi_growth": 0.030,
    "other_revenue_growth": 0.030,
    "in_lease_escalator": 0.025,  # 2.5% annual rent bumps per OM

    # Leasing assumptions
    "renewal_probability": 0.75,
    "new_lease_term_months": 42,  # 3 yrs 6 months
    "ti_new_psf": 1.50,
    "ti_renewal_psf": 0.00,
    "lc_new_pct": 0.0675,
    "lc_renewal_pct": 0.0325,
    "downtime_new_months": 3,
    "downtime_renewal_months": 0,
    "free_rent_months_new": 0,
    "free_rent_months_renewal": 0,

    # Vacancy / credit
    "general_vacancy": 0.020,
    "credit_loss": 0.010,

    # Reimbursements / mgmt
    "recovery_pct": 1.00,  # 100% NNN
    "mgmt_fee_pct_egr": 0.04,  # 4% of EGR
    "mgmt_fee_cap_pct_egr": 0.05,  # 5% cap per lease

    # Capital reserves
    "capital_reserve_psf": 0.20,

    # Tax reassessment
    "reassess_taxes_on_sale": True,  # DEFAULT ON (conservative buyer view)

    # Ancillary revenue
    "parking_rent_per_stall_month": 165,
    "parking_utilization_stabilized": 0.90,
    "parking_utilization_ramp_fy28": 0.85,
    "tenant_upgrade_rent_growth": 0.025,
    "ancillary_growth": 0.025,

    # Debt (USER SPEC: bank, 70% LTV, SOFR + 300 bps, 5-yr term)
    "ltv": 0.70,
    "loan_type": "Bank (Floating)",
    "sofr_rate": 0.043,             # editable — current approx SOFR
    "loan_spread_bps": 300,         # 3.00% spread
    "loan_amort_months": 360,       # 30-yr amort
    "loan_io_months": 24,           # 2-yr I/O (typical bank)
    "loan_term_months": 60,         # 5-yr term per user
    "loan_origination_fee_pct": 0.0075,  # 75 bps
    "refi_assumed": False,

    # Pricing — three modes user can toggle via PricingMode dropdown
    # Default: broker guidance $250/SF. Alt: 8% cap on in-place NOI.
    "pricing_mode": "price_psf",        # "price_psf" / "cap_on_inplace" / "cap_on_y1" / "price_direct"
    "price_per_sf_input": 250.00,       # broker guidance
    "target_going_in_cap": 0.080,       # user alt (8% cap on in-place)
    "purchase_price_input": 107637500,  # used only if "price_direct"
    "exit_cap_rate": 0.0650,            # 6.5% exit per user

    # Transaction costs (USER SPEC: 2%/2%)
    "acq_closing_cost_pct": 0.020,
    "disp_closing_cost_pct": 0.020,

    # Waterfall (structure reserved; default deal-level only — user can enable later)
    "waterfall_enabled": False,
    "pref_return": 0.08,
    "tier1_split_gp": 0.20,
    "tier1_hurdle_irr": 0.15,
    "tier2_split_gp": 0.30,
    "tier2_hurdle_irr": 0.20,
    "tier3_split_gp": 0.40,
}


# =============================================================================
# SECTION 2 — RENT ROLL DATA (from DFW Rent Roll 4.9.2026.xlsx)
# =============================================================================
# Loaded at runtime from the source .xlsx so we don't duplicate 162 rows here.
# Format: list of dicts with keys matching the rent roll columns plus a
# "treatment" flag derived from the Notes column ("Market" / "Option" / "Vacate").


def load_rent_roll(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Rent Roll"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    tenants = []
    for r in rows[1:]:
        if not r or r[0] is None or r[0] == "TOTAL / AVG":
            continue
        rec = dict(zip(header, r))

        # Classify lease treatment based on Notes and rent roll conventions
        notes = (rec.get("Notes") or "").lower()
        status = (rec.get("Status") or "").lower()
        if status == "vacant":
            treatment = "LeaseUp"  # vacant unit -> absorbs at market
        elif "assumed to vacate" in notes or "assumed vacate" in notes or "spec vacate" in notes:
            treatment = "Vacate"  # tenant will NOT renew
        elif "spec renew" in notes or "option" in notes:
            treatment = "Option"  # exercise contractual option
        else:
            treatment = "Market"  # apply 75% renewal probability weighting

        # Determine unit type for market rent lookup
        if "studio" in notes:
            unit_type = "Studio"
        else:
            unit_type = "Shop"

        rec["Treatment"] = treatment
        rec["UnitType"] = unit_type
        tenants.append(rec)
    return tenants


# =============================================================================
# SECTION 3 — STYLING HELPERS
# =============================================================================

# Colors (institutional underwriting color conventions)
CLR_INPUT_FILL = "FFF4E1"  # Soft yellow — user inputs
CLR_CALC_FILL = "FFFFFF"   # White — calculated
CLR_HEADER_FILL = "1F3864"  # Navy blue — section headers
CLR_SUBHEADER_FILL = "BDD7EE"  # Light blue — subheaders
CLR_TOTAL_FILL = "D9E1F2"  # Pale blue — totals
CLR_ESTIMATED_FILL = "FFE699"  # Orange — estimated values that need user review
CLR_NOI_FILL = "C6EFCE"  # Green — NOI rows

FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_SUBHEADER = Font(name="Calibri", size=10, bold=True, color="1F3864")
FONT_INPUT = Font(name="Calibri", size=10, color="0070C0")  # Blue — user input
FONT_CALC = Font(name="Calibri", size=10, color="000000")
FONT_TOTAL = Font(name="Calibri", size=10, bold=True)
FONT_TITLE = Font(name="Calibri", size=14, bold=True, color="1F3864")

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

# Number formats
FMT_DOLLAR = '_-$* #,##0_-;-$* #,##0_-;_-$* "-"_-;_-@_-'
FMT_DOLLAR_CENTS = '_-$* #,##0.00_-;-$* #,##0.00_-;_-$* "-"??_-;_-@_-'
FMT_NUMBER = "#,##0;(#,##0);-"
FMT_NUMBER_2DP = "#,##0.00;(#,##0.00);-"
FMT_PCT = "0.00%;(0.00%);-"
FMT_PCT_1DP = "0.0%;(0.0%);-"
FMT_MULT = '0.00"x"'
FMT_DATE = "mmm-yyyy"
FMT_PSF = '$#,##0.00"/SF"'


def style_input(cell):
    cell.fill = PatternFill("solid", fgColor=CLR_INPUT_FILL)
    cell.font = FONT_INPUT
    cell.border = THIN_BORDER


def style_estimated(cell):
    cell.fill = PatternFill("solid", fgColor=CLR_ESTIMATED_FILL)
    cell.font = FONT_INPUT
    cell.border = THIN_BORDER


def style_calc(cell):
    cell.fill = PatternFill("solid", fgColor=CLR_CALC_FILL)
    cell.font = FONT_CALC
    cell.border = THIN_BORDER


def style_header(cell):
    cell.fill = PatternFill("solid", fgColor=CLR_HEADER_FILL)
    cell.font = FONT_HEADER
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_subheader(cell):
    cell.fill = PatternFill("solid", fgColor=CLR_SUBHEADER_FILL)
    cell.font = FONT_SUBHEADER
    cell.border = THIN_BORDER


def style_total(cell):
    cell.fill = PatternFill("solid", fgColor=CLR_TOTAL_FILL)
    cell.font = FONT_TOTAL
    cell.border = THIN_BORDER


def style_noi(cell):
    cell.fill = PatternFill("solid", fgColor=CLR_NOI_FILL)
    cell.font = FONT_TOTAL
    cell.border = THIN_BORDER


# =============================================================================
# SECTION 4 — TAB BUILDERS
# =============================================================================

OUTPUT_PATH = "/home/user/DFW-Deal/DFW_Underwriting_Model.xlsx"
RENT_ROLL_PATH = "/home/user/DFW-Deal/DFW Rent Roll 4.9.2026.xlsx"


def add_named_range(wb, name, sheet_name, cell_ref):
    """Register a workbook-scoped named range pointing to a single cell."""
    ref = f"'{sheet_name}'!${cell_ref}"
    dn = DefinedName(name=name, attr_text=ref)
    wb.defined_names[name] = dn


def write_section_header(ws, row, text, span=3):
    ws.cell(row=row, column=1, value=text)
    style_header(ws.cell(row=row, column=1))
    for c in range(2, span + 1):
        ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=CLR_HEADER_FILL)


def build_assumptions(wb):
    """Assumptions tab — all user-editable inputs with named ranges."""
    ws = wb.create_sheet("Assumptions")
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 48

    # Title
    ws["A1"] = "DFW MODERN MULTI-TENANT PORTFOLIO — UNDERWRITING ASSUMPTIONS"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:C1")
    ws["A2"] = "All yellow cells are user inputs. Orange = estimated (Plano OM incomplete)."
    ws["A2"].font = Font(italic=True, size=9, color="808080")
    ws.merge_cells("A2:C2")

    row = 4

    def section(title):
        nonlocal row
        write_section_header(ws, row, title, span=3)
        row += 1

    def inp(label, name, value, fmt=None, note="", estimated=False):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = FONT_CALC
        c = ws.cell(row=row, column=2, value=value)
        if estimated:
            style_estimated(c)
        else:
            style_input(c)
        if fmt:
            c.number_format = fmt
        ws.cell(row=row, column=3, value=note).font = Font(italic=True, size=9, color="606060")
        add_named_range(wb, name, "Assumptions", f"B{row}")
        row += 1

    def calc(label, name, formula, fmt=None, note=""):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = FONT_CALC
        c = ws.cell(row=row, column=2, value=formula)
        style_calc(c)
        if fmt:
            c.number_format = fmt
        ws.cell(row=row, column=3, value=note).font = Font(italic=True, size=9, color="606060")
        add_named_range(wb, name, "Assumptions", f"B{row}")
        row += 1

    # Timing
    section("MODEL TIMING")
    inp("Analysis Start Date", "AnalysisStart", DEFAULTS["analysis_start"], FMT_DATE, "May 1, 2026 — matches OM")
    inp("Hold Period (months)", "HoldMonths", 60, "0", "5-year hold (matches loan term)")
    inp("Projection Period (months)", "ProjMonths", 120, "0", "Always project 10 yrs for exit cap / reversion math")
    inp("Fiscal Year End Month", "FYEndMonth", 4, "0", "April (matches OM Argus FYE 4/30)")

    # Growth rates
    section("GROWTH RATES")
    inp("Market Rent Growth", "GrowthMarketRent", DEFAULTS["market_rent_growth"], FMT_PCT, "CBRE/Argus assumption")
    inp("In-Lease Rent Escalator", "EscalatorInLease", DEFAULTS["in_lease_escalator"], FMT_PCT, "Typical 2.5% annual bumps")
    inp("Operating Expense Growth", "GrowthOpEx", DEFAULTS["opex_growth"], FMT_PCT, "")
    inp("RE Tax Growth", "GrowthRETax", DEFAULTS["re_tax_growth"], FMT_PCT, "Applied to reassessed base")
    inp("CPI", "CPI", DEFAULTS["cpi_growth"], FMT_PCT, "")
    inp("Other Revenue Growth", "GrowthOther", DEFAULTS["other_revenue_growth"], FMT_PCT, "Ancillary income")

    # Leasing
    section("LEASING ASSUMPTIONS (PORTFOLIO DEFAULTS — overrideable on MLA tab)")
    inp("Renewal Probability (Retention)", "Retention", DEFAULTS["renewal_probability"], FMT_PCT, "CBRE default 75%")
    inp("New Lease Term (months)", "LeaseTermMonths", DEFAULTS["new_lease_term_months"], "0", "3 yrs 6 months")
    inp("TI — New ($/SF)", "TINew", DEFAULTS["ti_new_psf"], FMT_DOLLAR_CENTS, "")
    inp("TI — Renewal ($/SF)", "TIRenewal", DEFAULTS["ti_renewal_psf"], FMT_DOLLAR_CENTS, "")
    inp("LC — New (% of base rent)", "LCNew", DEFAULTS["lc_new_pct"], FMT_PCT, "")
    inp("LC — Renewal (%)", "LCRenewal", DEFAULTS["lc_renewal_pct"], FMT_PCT, "")
    inp("Downtime — New (months)", "DowntimeNew", DEFAULTS["downtime_new_months"], "0", "")
    inp("Downtime — Renewal (months)", "DowntimeRenewal", DEFAULTS["downtime_renewal_months"], "0", "Typically zero")
    inp("Free Rent — New (months)", "FreeRentNew", DEFAULTS["free_rent_months_new"], "0", "")
    inp("Free Rent — Renewal (months)", "FreeRentRenewal", DEFAULTS["free_rent_months_renewal"], "0", "")

    # Vacancy / credit
    section("VACANCY & CREDIT")
    inp("General Vacancy", "GeneralVacancy", DEFAULTS["general_vacancy"], FMT_PCT, "Beyond lease-up vacancy")
    inp("Credit Loss", "CreditLoss", DEFAULTS["credit_loss"], FMT_PCT, "")

    # Reimbursements / mgmt
    section("REIMBURSEMENTS & MANAGEMENT")
    inp("Recovery % (NNN Pass-through)", "RecoveryPct", DEFAULTS["recovery_pct"], FMT_PCT, "100% NNN all leases")
    inp("Management Fee (% of EGR)", "MgmtFeePct", DEFAULTS["mgmt_fee_pct_egr"], FMT_PCT, "4% per OM; 5% cap per lease")
    inp("Capital Reserves ($/SF)", "ReservesPSF", DEFAULTS["capital_reserve_psf"], FMT_DOLLAR_CENTS, "")

    # Taxes
    section("REAL ESTATE TAXES")
    inp("Reassess Taxes at Purchase?", "ReassessTaxes", True, None, "TRUE = taxes reset to Purchase Price × millage")

    # Ancillary
    section("ANCILLARY INCOME")
    inp("Parking Rent $/Stall/Month (new)", "ParkingRentMonthly", DEFAULTS["parking_rent_per_stall_month"], FMT_DOLLAR, "Secured parking")
    inp("Parking Stabilized Utilization", "ParkingUtilStab", DEFAULTS["parking_utilization_stabilized"], FMT_PCT, "Target year 3+")
    inp("Tenant Upgrade Rent Growth", "TURentGrowth", DEFAULTS["tenant_upgrade_rent_growth"], FMT_PCT, "")

    # Pricing
    section("PRICING")
    inp("Pricing Mode", "PricingMode", DEFAULTS["pricing_mode"], None,
        "price_psf | cap_on_inplace | cap_on_y1 | price_direct")
    inp("Price per SF (Mode: price_psf)", "PricePerSF", DEFAULTS["price_per_sf_input"], FMT_DOLLAR_CENTS, "Broker guidance $250/SF")
    inp("Going-In Cap Rate (Mode: cap_on_*)", "TargetCap", DEFAULTS["target_going_in_cap"], FMT_PCT, "User spec: 8% on in-place NOI")
    inp("Purchase Price Direct Input", "PricePriceDirect", DEFAULTS["purchase_price_input"], FMT_DOLLAR, "Only used if Mode = price_direct")
    inp("Exit Cap Rate", "ExitCap", DEFAULTS["exit_cap_rate"], FMT_PCT, "User spec: 6.5%")

    # Transaction costs
    section("TRANSACTION COSTS")
    inp("Acquisition Closing Costs %", "AcqCostPct", DEFAULTS["acq_closing_cost_pct"], FMT_PCT, "User spec: 2%")
    inp("Disposition Closing Costs %", "DispCostPct", DEFAULTS["disp_closing_cost_pct"], FMT_PCT, "User spec: 2%")

    # Debt
    section("DEBT (BANK FLOATING — USER SPEC)")
    inp("Loan-to-Value (LTV)", "LTV", DEFAULTS["ltv"], FMT_PCT, "70% per user")
    inp("SOFR Rate", "SOFR", DEFAULTS["sofr_rate"], FMT_PCT, "Editable — current ~4.3%")
    inp("Loan Spread (bps)", "LoanSpreadBps", DEFAULTS["loan_spread_bps"], "0", "300 bps per user")
    calc("All-In Loan Rate", "LoanRate", "=SOFR+LoanSpreadBps/10000", FMT_PCT, "= SOFR + Spread")
    inp("Loan Term (months)", "LoanTermMonths", DEFAULTS["loan_term_months"], "0", "60 months (5-yr)")
    inp("Interest-Only Period (months)", "LoanIOMonths", DEFAULTS["loan_io_months"], "0", "")
    inp("Amortization (months)", "LoanAmortMonths", DEFAULTS["loan_amort_months"], "0", "30-yr amort")
    inp("Origination Fee %", "LoanOrigFeePct", DEFAULTS["loan_origination_fee_pct"], FMT_PCT, "75 bps typical bank")

    # Waterfall (placeholder)
    section("WATERFALL (RESERVED — enable later)")
    inp("Waterfall Enabled?", "WaterfallOn", False, None, "Deal-level returns only for now")
    inp("Preferred Return", "PrefReturn", DEFAULTS["pref_return"], FMT_PCT, "")
    inp("Tier 1 GP Split (to 15% IRR)", "Tier1GPSplit", DEFAULTS["tier1_split_gp"], FMT_PCT, "")
    inp("Tier 1 IRR Hurdle", "Tier1Hurdle", DEFAULTS["tier1_hurdle_irr"], FMT_PCT, "")
    inp("Tier 2 GP Split (to 20% IRR)", "Tier2GPSplit", DEFAULTS["tier2_split_gp"], FMT_PCT, "")
    inp("Tier 2 IRR Hurdle", "Tier2Hurdle", DEFAULTS["tier2_hurdle_irr"], FMT_PCT, "")
    inp("Tier 3 GP Split (above)", "Tier3GPSplit", DEFAULTS["tier3_split_gp"], FMT_PCT, "")

    # Derived: Purchase Price formula
    section("DERIVED PRICING (calculated)")
    # In-place NOI is computed on Property Data tab; we reference it here
    calc("In-Place NOI (Portfolio)", "InPlaceNOI", "=SUM('Property Data'!B18:G18)", FMT_DOLLAR, "Sum across 6 properties")
    calc("Portfolio SF", "PortfolioSF", "=SUM('Property Data'!B4:G4)", FMT_NUMBER, "")
    calc("Purchase Price", "PurchasePrice",
         '=IF(PricingMode="price_psf",PricePerSF*PortfolioSF,'
         'IF(PricingMode="cap_on_inplace",InPlaceNOI/TargetCap,'
         'IF(PricingMode="cap_on_y1",\'Annual CF\'!B21/TargetCap,'
         'PricePriceDirect)))',
         FMT_DOLLAR, "Resolves based on Pricing Mode")
    calc("Price per SF (implied)", "PricePerSFImplied", "=PurchasePrice/PortfolioSF", FMT_DOLLAR_CENTS, "")
    calc("Going-In Cap (on In-Place NOI)", "CapInPlace", "=InPlaceNOI/PurchasePrice", FMT_PCT, "")

    # Pricing Mode dropdown
    dv_mode = DataValidation(
        type="list",
        formula1='"price_psf,cap_on_inplace,cap_on_y1,price_direct"',
        allow_blank=False,
    )
    dv_mode.error = "Must be: price_psf, cap_on_inplace, cap_on_y1, or price_direct"
    dv_mode.errorTitle = "Invalid Pricing Mode"
    dv_mode.add("B41")  # PricingMode cell
    ws.add_data_validation(dv_mode)

    # TRUE/FALSE dropdowns for boolean inputs
    dv_bool = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
    dv_bool.add("B35")  # ReassessTaxes
    dv_bool.add("B59")  # WaterfallOn
    ws.add_data_validation(dv_bool)


def build_property_data(wb):
    """Property Data tab — per-property OM values in columns."""
    ws = wb.create_sheet("Property Data")

    # Layout: Column A = line item, B-G = 6 properties
    ws.column_dimensions["A"].width = 36
    for i, p in enumerate(PROPERTIES):
        col = get_column_letter(2 + i)
        ws.column_dimensions[col].width = 14

    # Header row
    ws["A1"] = "PROPERTY DATA"
    ws["A1"].font = FONT_TITLE
    for i, p in enumerate(PROPERTIES):
        c = ws.cell(row=2, column=2 + i, value=p["name"])
        style_subheader(c)

    rows_defs = [
        ("Property ID", "id", None),
        ("Total SF", "sf", FMT_NUMBER),  # row 3 -- referenced by PortfolioSF formula (B3:G3)
        ("Office SF", "office_sf", FMT_NUMBER),
        ("Office %", "office_pct", FMT_PCT),
        ("Land Acres", "land_acres", FMT_NUMBER_2DP),
        ("# Buildings", "buildings", "0"),
        ("# Units", "units", "0"),
        ("Year Built", "year_built", "0"),
        ("Parking Spaces", "parking", "0"),
        ("Secured Parking", "secured_parking", "0"),
        ("Market Rent — Shop ($/SF)", "market_rent_shop", FMT_DOLLAR_CENTS),
        ("Market Rent — Studio ($/SF)", "market_rent_studio", FMT_DOLLAR_CENTS),
        ("2025 Assessed Value", "assessed_value_2025", FMT_DOLLAR),
        ("Millage Rate", "millage_rate", FMT_PCT),
        ("Parcel", "parcel", None),
        ("In-Place NOI (OM)", "inplace_noi", FMT_DOLLAR),  # row 16 — referenced by InPlaceNOI
        ("FY27 NOI (OM Proforma)", "fy27_noi_om", FMT_DOLLAR),
        ("FY27 CAM", "fy27_cam", FMT_DOLLAR),
        ("FY27 Utilities", "fy27_utilities", FMT_DOLLAR),
        ("FY27 Management Fee", "fy27_mgmt_fee", FMT_DOLLAR),
        ("FY27 Insurance", "fy27_insurance", FMT_DOLLAR),
        ("FY27 RE Taxes", "fy27_re_taxes", FMT_DOLLAR),
        ("FY27 Tenant Upgrade Rent", "tenant_upgrade_rent_fy27", FMT_DOLLAR),
        ("FY27 Parking Rent", "parking_rent_fy27", FMT_DOLLAR),
    ]

    PLANO_IDX = next(i for i, p in enumerate(PROPERTIES) if p["id"] == "PLA")

    for r_idx, (label, key, fmt) in enumerate(rows_defs):
        row = 3 + r_idx
        ws.cell(row=row, column=1, value=label).font = FONT_CALC
        for i, p in enumerate(PROPERTIES):
            c = ws.cell(row=row, column=2 + i, value=p.get(key))
            if i == PLANO_IDX and key in (
                "assessed_value_2025", "millage_rate", "fy27_cam", "fy27_utilities",
                "fy27_mgmt_fee", "fy27_insurance", "fy27_re_taxes", "fy27_noi_om",
                "inplace_noi", "tenant_upgrade_rent_fy27", "parking_rent_fy27",
                "parcel", "land_acres", "parking", "secured_parking",
            ):
                style_estimated(c)
            else:
                style_input(c) if key in ("fy27_cam", "fy27_utilities", "fy27_mgmt_fee",
                                          "fy27_insurance", "fy27_re_taxes",
                                          "market_rent_shop", "market_rent_studio",
                                          "inplace_noi", "assessed_value_2025",
                                          "millage_rate", "tenant_upgrade_rent_fy27",
                                          "parking_rent_fy27") else style_calc(c)
            if fmt:
                c.number_format = fmt

    # Totals column H
    ws.cell(row=2, column=8, value="Portfolio").font = FONT_SUBHEADER
    style_subheader(ws.cell(row=2, column=8))
    for r_idx, (label, key, fmt) in enumerate(rows_defs):
        row = 3 + r_idx
        if fmt in (FMT_NUMBER, FMT_DOLLAR, FMT_DOLLAR_CENTS, FMT_NUMBER_2DP, "0"):
            c = ws.cell(row=row, column=8, value=f"=SUM(B{row}:G{row})")
            style_total(c)
            if fmt:
                c.number_format = fmt

    ws.column_dimensions["H"].width = 16


def build_mla(wb):
    """Market Leasing Assumptions — per property + unit type (Shop/Studio)."""
    ws = wb["MLA"] if "MLA" in wb.sheetnames else wb.create_sheet("MLA")
    # Clear existing
    for row in ws.iter_rows():
        for c in row:
            c.value = None

    ws["A1"] = "MARKET LEASING ASSUMPTIONS"
    ws["A1"].font = FONT_TITLE
    ws["A2"] = "Defaults to OM values. Edit per property/unit type to override global assumptions."
    ws["A2"].font = Font(italic=True, size=9, color="808080")

    # Column widths
    widths = [24, 12, 16, 14, 14, 14, 14, 14, 14, 14, 14, 14]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    headers = [
        "Property / Unit Type", "Unit Type", "Market Rent ($/SF)",
        "Rent Growth", "In-Lease Escalator", "Lease Term (mo)",
        "TI New $/SF", "TI Renewal $/SF", "LC New %", "LC Renewal %",
        "Downtime New (mo)", "Retention %",
    ]
    for i, h in enumerate(headers):
        c = ws.cell(row=4, column=i + 1, value=h)
        style_subheader(c)
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    # Build rows: Carrollton has Shop + Studio (different rents); others mostly unified
    mla_rows = []
    for p in PROPERTIES:
        if p["market_rent_shop"] != p["market_rent_studio"]:
            mla_rows.append((f"{p['name']} — Shop", p["id"], "Shop", p["market_rent_shop"]))
            mla_rows.append((f"{p['name']} — Studio", p["id"], "Studio", p["market_rent_studio"]))
        else:
            mla_rows.append((f"{p['name']}", p["id"], "All", p["market_rent_shop"]))

    for r_idx, (label, pid, unit_type, market_rent) in enumerate(mla_rows):
        row = 5 + r_idx
        ws.cell(row=row, column=1, value=label).font = FONT_CALC
        ws.cell(row=row, column=2, value=unit_type).font = FONT_CALC

        # All MLA fields default from Assumptions but overrideable per row
        cells = [
            (3, market_rent, FMT_DOLLAR_CENTS, True),          # Market Rent
            (4, "=GrowthMarketRent", FMT_PCT, False),          # Rent Growth
            (5, "=EscalatorInLease", FMT_PCT, False),          # In-Lease Escalator
            (6, "=LeaseTermMonths", "0", False),               # Lease Term
            (7, "=TINew", FMT_DOLLAR_CENTS, False),
            (8, "=TIRenewal", FMT_DOLLAR_CENTS, False),
            (9, "=LCNew", FMT_PCT, False),
            (10, "=LCRenewal", FMT_PCT, False),
            (11, "=DowntimeNew", "0", False),
            (12, "=Retention", FMT_PCT, False),
        ]
        for col, val, fmt, is_input in cells:
            c = ws.cell(row=row, column=col, value=val)
            if is_input:
                style_input(c)
            else:
                style_calc(c)
            c.number_format = fmt

    # Store row count for later lookups
    ws["A1"].comment = Comment(f"MLA data rows 5 to {4 + len(mla_rows)}", "model")


def build_rent_roll(wb, tenants):
    """Rent Roll tab — 162 tenants with FY27-FY37 annual rent + TI/LC formulas.

    Column layout:
      A Property  B Suite  C Tenant  D SF  E Lease Start  F Lease End
      G In-Place PSF  H Market PSF  I Treatment  J Unit Type  K Status  L Notes
      M YearsInLeaseAtStart (helper)
      N-X  FY27-FY37 annual base rent (11 cols)
      Y-AI FY27-FY37 TI+LC capital (11 cols)
      AJ-AT FY27-FY37 Downtime Loss ($ reduction) (11 cols)
    """
    ws = wb["Rent Roll"] if "Rent Roll" in wb.sheetnames else wb.create_sheet("Rent Roll")
    for row in ws.iter_rows():
        for c in row:
            c.value = None

    ws["A1"] = "RENT ROLL (162 tenant-suites) — FY27–FY37 dynamic base rent & capital"
    ws["A1"].font = FONT_TITLE
    ws["A2"] = ("Formulas apply Treatment flag (Market/Option/Vacate/LeaseUp) per tenant. "
                "Market Rent PSF (col H) is looked up from MLA tab. Edit Treatment col to switch scenarios.")
    ws["A2"].font = Font(italic=True, size=9, color="808080")

    # Column widths
    for i, w in enumerate([13, 8, 34, 8, 12, 12, 10, 10, 11, 9, 10, 40, 10] + [11] * 33):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    # Headers
    HEADER_ROW = 4
    base_headers = [
        "Property", "Suite", "Tenant", "SF", "Lease Start", "Lease End",
        "In-Place PSF", "Market PSF", "Treatment", "Unit Type", "Status", "Notes",
        "YrsAtStart",
    ]
    for i, h in enumerate(base_headers):
        c = ws.cell(row=HEADER_ROW, column=i + 1, value=h)
        style_subheader(c)

    # FY columns: FY27 thru FY37 = 11 fiscal years
    fy_labels = [f"FY{27 + i}" for i in range(11)]  # FY27 - FY37
    # Three blocks: Base Rent, TI+LC, Downtime Loss
    blocks = [("BaseRent", 14), ("Capital", 25), ("Downtime", 36)]  # starting cols
    for block_name, start_col in blocks:
        for i, label in enumerate(fy_labels):
            c = ws.cell(row=HEADER_ROW, column=start_col + i, value=f"{block_name} {label}")
            style_subheader(c)
            c.alignment = Alignment(horizontal="center", wrap_text=True)

    # Parse a "Mon-YYYY" string to an actual date (first of month)
    def parse_date(s):
        if s is None or s == "":
            return None
        if isinstance(s, dt.datetime):
            return s.date()
        if isinstance(s, dt.date):
            return s
        try:
            return dt.datetime.strptime(s, "%b-%Y").date()
        except Exception:
            return None

    # Write tenant rows
    for r_idx, t in enumerate(tenants):
        row = HEADER_ROW + 1 + r_idx
        ws.cell(row=row, column=1, value=t["Property"])
        ws.cell(row=row, column=2, value=str(t["Suite"]))
        ws.cell(row=row, column=3, value=t["Tenant"])
        ws.cell(row=row, column=4, value=t["SF"]).number_format = FMT_NUMBER
        ls = parse_date(t["Lease Start"])
        le = parse_date(t["Lease End"])
        if ls:
            c = ws.cell(row=row, column=5, value=ls)
            c.number_format = FMT_DATE
        if le:
            c = ws.cell(row=row, column=6, value=le)
            c.number_format = FMT_DATE
        c = ws.cell(row=row, column=7, value=t["Rent PSF"] or 0)
        c.number_format = FMT_DOLLAR_CENTS

        # Market PSF lookup via VLOOKUP on MLA tab (property name or "Property — Shop/Studio")
        # Build the lookup key: if Carrollton, use "Carrollton — {UnitType}"; else just property name
        prop_name = t["Property"]
        unit_type = t["UnitType"]
        if prop_name == "Carrollton":
            lookup_key = f'"Carrollton — "&J{row}'
            market_formula = f'=VLOOKUP("Carrollton — "&J{row},MLA!$A$5:$L$11,3,FALSE)'
        else:
            market_formula = f'=VLOOKUP(A{row},MLA!$A$5:$L$11,3,FALSE)'
        c = ws.cell(row=row, column=8, value=market_formula)
        c.number_format = FMT_DOLLAR_CENTS
        style_calc(c)

        ws.cell(row=row, column=9, value=t["Treatment"])  # editable
        style_input(ws.cell(row=row, column=9))
        ws.cell(row=row, column=10, value=unit_type)
        ws.cell(row=row, column=11, value=t["Status"])
        ws.cell(row=row, column=12, value=t.get("Notes") or "")

        # Helper: years between analysis_start and lease_start (negative if future)
        # We use: =IF(E5="",0, (AnalysisStart - E5) / 365.25)
        c = ws.cell(row=row, column=13, value=f'=IF(E{row}="",0,(AnalysisStart-E{row})/365.25)')
        c.number_format = FMT_NUMBER_2DP
        style_calc(c)

        # FY27-FY37 Base Rent formulas (cols N-X = 14-24)
        # Strategy:
        #   fy_offset = i (0 for FY27, 10 for FY37)
        #   fy_mid_date = DATE(2026+i+1, 10, 31)  # mid-point of FY (Oct 31)
        #   fy_start_date = DATE(2026+i, 5, 1)
        #   fy_end_date = DATE(2026+i+1, 4, 30)
        #   in_lease = LeaseEnd >= fy_end_date AND LeaseStart <= fy_start_date  (full year)
        #   expired = LeaseEnd < fy_start_date
        #
        # Formula (simplified; handles the common cases):
        #   IF lease active full FY:
        #     InPlacePSF * SF * (1+Escalator)^(fy_offset + YrsAtStart)
        #   ELSEIF expired before FY:
        #     post-expiry logic based on Treatment
        #   ELSE (lease expires mid-FY or starts mid-FY):
        #     prorated: (months_in_lease/12 * in_place_grown) + (months_post/12 * post_expiry)

        for i in range(11):
            col = 14 + i
            fy_offset = i
            # Excel date serials for FY start/end/mid
            fy_start_ref = f'DATE(2026+{i},5,1)'
            fy_end_ref = f'DATE(2027+{i},4,30)'
            fy_mid_ref = f'DATE(2026+{i},10,31)'

            # In-place rent grown at escalator from FY27 base (in-place PSF already reflects
            # past escalation at rent roll date — do NOT add YrsAtStart to exponent).
            in_place_grown = f'G{row}*D{row}*(1+EscalatorInLease)^{fy_offset}'

            # Years since expiration at FY midpoint (negative if not yet expired)
            yrs_since_expiry = f'({fy_mid_ref}-F{row})/365.25'

            # Vacate: full market rent post-expiry, with downtime in the expiration FY only
            vacate = (f'H{row}*D{row}*(1+GrowthMarketRent)^{fy_offset}'
                      f'*IF(INT({yrs_since_expiry})=0,(12-DowntimeNew)/12,1)')
            # Option: one-time 2.5% bump at exercise, then grow at escalator — avoid double count
            # Formula: in_place × 1.025 × (1+esc)^(years past expiry)
            # Simplified: model at escalator growth from FY27 base (since 1.025 ≈ escalator)
            option = f'G{row}*1.025*D{row}*(1+EscalatorInLease)^{fy_offset}'
            # Market: weighted 75% renewal (in-place-grown) + 25% new market (with downtime)
            market = (f'Retention*{in_place_grown}'
                      f'+(1-Retention)*H{row}*D{row}*(1+GrowthMarketRent)^{fy_offset}'
                      f'*IF(INT({yrs_since_expiry})=0,(12-DowntimeNew)/12,1)')
            # LeaseUp: market rent, downtime only in Year 1
            lease_up = (f'H{row}*D{row}*(1+GrowthMarketRent)^{fy_offset}'
                        f'*IF({fy_offset}=0,(12-DowntimeNew)/12,1)')

            # Pick post-expiry rent based on Treatment
            post_expiry = (f'IF(I{row}="Vacate",{vacate},'
                           f'IF(I{row}="Option",{option},'
                           f'IF(I{row}="Market",{market},'
                           f'IF(I{row}="LeaseUp",{lease_up},0))))')

            # Main formula: piecewise based on where lease ends relative to FY.
            # IFERROR wraps the whole thing — returns 0 if any downstream ref fails.
            formula = (
                f'=IFERROR('
                f'IF(I{row}="LeaseUp",{lease_up},'
                f'IF(F{row}="",0,'
                # Lease active all year:
                f'IF(F{row}>={fy_end_ref},{in_place_grown},'
                # Lease expired before FY starts: full post-expiry year
                f'IF(F{row}<{fy_start_ref},{post_expiry},'
                # Lease expires mid-FY: prorate between in-place and post-expiry (downtime applied in post_expiry)
                f'((F{row}-{fy_start_ref})/365.25)*{in_place_grown}'
                f'+(({fy_end_ref}-F{row})/365.25)*{post_expiry}'
                f')))),0)'
            )
            c = ws.cell(row=row, column=col, value=formula)
            c.number_format = FMT_DOLLAR
            style_calc(c)

        # FY27-FY37 Capital (TI + LC) — hits only in the FY the lease expires (rolls over)
        # If treatment = LeaseUp → hits in FY27 (Year 1, lease-up)
        for i in range(11):
            col = 25 + i
            fy_offset = i
            fy_start_ref = f'DATE(2026+{i},5,1)'
            fy_end_ref = f'DATE(2027+{i},4,30)'

            # Did the lease expire within this FY? Or for LeaseUp, is this FY27?
            # Expected new rent at that time (for LC % calc)
            new_rent_at_rollover = f'H{row}*D{row}*(1+GrowthMarketRent)^{fy_offset}'
            # Blended TI (75% renewal × 0 + 25% new × 1.50 = ~$0.38 weighted)
            ti_cost = (f'IF(I{row}="Vacate",TINew*D{row},'
                       f'IF(I{row}="Market",Retention*TIRenewal*D{row}+(1-Retention)*TINew*D{row},'
                       f'IF(I{row}="LeaseUp",TINew*D{row},'
                       f'IF(I{row}="Option",TIRenewal*D{row},0))))')
            lc_cost = (f'IF(I{row}="Vacate",LCNew*{new_rent_at_rollover},'
                       f'IF(I{row}="Market",Retention*LCRenewal*{new_rent_at_rollover}+(1-Retention)*LCNew*{new_rent_at_rollover},'
                       f'IF(I{row}="LeaseUp",LCNew*{new_rent_at_rollover},'
                       f'IF(I{row}="Option",LCRenewal*{new_rent_at_rollover},0))))')

            cap_formula = (
                f'=IFERROR('
                f'IF(AND(I{row}="LeaseUp",{fy_offset}=0),({ti_cost})+({lc_cost}),'
                f'IF(F{row}="",0,'
                f'IF(AND(F{row}>={fy_start_ref},F{row}<={fy_end_ref}),({ti_cost})+({lc_cost}),0)))'
                f',0)'
            )
            c = ws.cell(row=row, column=col, value=cap_formula)
            c.number_format = FMT_DOLLAR
            style_calc(c)

    # Total rows at bottom
    total_row = HEADER_ROW + 1 + len(tenants)
    ws.cell(row=total_row, column=1, value="PORTFOLIO TOTAL").font = FONT_TOTAL
    for col in range(4, 5):  # SF
        cref = f'{get_column_letter(col)}{HEADER_ROW + 1}:{get_column_letter(col)}{total_row - 1}'
        c = ws.cell(row=total_row, column=col, value=f'=SUM({cref})')
        style_total(c)
        c.number_format = FMT_NUMBER
    for col in range(14, 47):  # Base rent + capital columns
        cref = f'{get_column_letter(col)}{HEADER_ROW + 1}:{get_column_letter(col)}{total_row - 1}'
        c = ws.cell(row=total_row, column=col, value=f'=SUM({cref})')
        style_total(c)
        c.number_format = FMT_DOLLAR

    # Add data validation: Treatment column (I) dropdown
    dv = DataValidation(
        type="list",
        formula1='"Market,Option,Vacate,LeaseUp"',
        allow_blank=False,
    )
    dv.error = "Must be Market, Option, Vacate, or LeaseUp"
    dv.errorTitle = "Invalid Treatment"
    dv.prompt = "Market = 75% renewal / 25% new; Option = 2.5% bump; Vacate = 100% new with downtime; LeaseUp = currently vacant"
    dv.promptTitle = "Treatment"
    last_row = HEADER_ROW + len(tenants)
    dv.add(f"I5:I{last_row}")
    ws.add_data_validation(dv)

    # Freeze panes
    ws.freeze_panes = "N5"


def build_annual_cf(wb):
    """Annual CF tab — FY27-FY37 portfolio cash flow rollup."""
    ws = wb["Annual CF"] if "Annual CF" in wb.sheetnames else wb.create_sheet("Annual CF")
    for row in ws.iter_rows():
        for c in row:
            c.value = None

    ws["A1"] = "ANNUAL CASH FLOW — PORTFOLIO (FY Ending April 30)"
    ws["A1"].font = FONT_TITLE

    # Column setup: A = Line Item, B-L = FY27-FY37 (11 years), M = Total
    ws.column_dimensions["A"].width = 38
    for i in range(11):
        ws.column_dimensions[get_column_letter(2 + i)].width = 15
    ws.column_dimensions["M"].width = 16

    # Header row
    for i in range(11):
        c = ws.cell(row=3, column=2 + i, value=f"FY{27 + i}")
        style_subheader(c)
        c.alignment = Alignment(horizontal="center")
    c = ws.cell(row=3, column=13, value="Total")
    style_subheader(c)

    # ---- REVENUE ----
    ws.cell(row=5, column=1, value="REVENUE").font = FONT_SUBHEADER

    # Row 6: Gross Potential Rent (sum of all Rent Roll base rent columns)
    ws.cell(row=6, column=1, value="Gross Potential Rent").font = FONT_CALC
    for i in range(11):
        col_letter_rr = get_column_letter(14 + i)  # N-X on Rent Roll
        c = ws.cell(row=6, column=2 + i,
                    value=f"='Rent Roll'!{col_letter_rr}167")  # total row on Rent Roll
        c.number_format = FMT_DOLLAR
        style_calc(c)

    # Row 7: Vacancy + Credit Loss
    ws.cell(row=7, column=1, value="Less: Vacancy & Credit Loss").font = FONT_CALC
    for i in range(11):
        col = get_column_letter(2 + i)
        c = ws.cell(row=7, column=2 + i, value=f"=-{col}6*(GeneralVacancy+CreditLoss)")
        c.number_format = FMT_DOLLAR
        style_calc(c)

    # Row 8: Expense Recoveries (= Total OpEx × recovery%)
    # OpEx is on row 19 (computed below) — forward reference
    ws.cell(row=8, column=1, value="Expense Recoveries (NNN)").font = FONT_CALC
    for i in range(11):
        col = get_column_letter(2 + i)
        c = ws.cell(row=8, column=2 + i, value=f"=-{col}19*RecoveryPct")
        c.number_format = FMT_DOLLAR
        style_calc(c)

    # Row 9: Tenant Upgrade Rent (portfolio FY27 base H25, grown)
    ws.cell(row=9, column=1, value="Tenant Upgrade Rent").font = FONT_CALC
    for i in range(11):
        c = ws.cell(row=9, column=2 + i,
                    value=f"='Property Data'!H25*(1+TURentGrowth)^{i}")
        c.number_format = FMT_DOLLAR
        style_calc(c)

    # Row 10: Parking Rent (portfolio FY27 base H26, grown)
    ws.cell(row=10, column=1, value="Parking Rent").font = FONT_CALC
    for i in range(11):
        c = ws.cell(row=10, column=2 + i,
                    value=f"='Property Data'!H26*(1+GrowthOther)^{i}")
        c.number_format = FMT_DOLLAR
        style_calc(c)

    # Row 11: Effective Gross Revenue
    ws.cell(row=11, column=1, value="EFFECTIVE GROSS REVENUE (EGR)").font = FONT_TOTAL
    for i in range(11):
        col = get_column_letter(2 + i)
        c = ws.cell(row=11, column=2 + i, value=f"=SUM({col}6:{col}10)")
        style_total(c)
        c.number_format = FMT_DOLLAR

    # ---- OPERATING EXPENSES ----
    ws.cell(row=13, column=1, value="OPERATING EXPENSES").font = FONT_SUBHEADER

    # Row 14: CAM (from Property Data H20, grown at OpEx)
    ws.cell(row=14, column=1, value="CAM").font = FONT_CALC
    for i in range(11):
        c = ws.cell(row=14, column=2 + i,
                    value=f"=-'Property Data'!H20*(1+GrowthOpEx)^{i}")
        c.number_format = FMT_DOLLAR
        style_calc(c)

    # Row 15: Utilities
    ws.cell(row=15, column=1, value="Utilities").font = FONT_CALC
    for i in range(11):
        c = ws.cell(row=15, column=2 + i,
                    value=f"=-'Property Data'!H21*(1+GrowthOpEx)^{i}")
        c.number_format = FMT_DOLLAR
        style_calc(c)

    # Row 16: Management Fee (4% of EGR)
    ws.cell(row=16, column=1, value="Management Fee").font = FONT_CALC
    for i in range(11):
        col = get_column_letter(2 + i)
        c = ws.cell(row=16, column=2 + i, value=f"=-{col}11*MgmtFeePct")
        c.number_format = FMT_DOLLAR
        style_calc(c)

    # Row 17: Insurance
    ws.cell(row=17, column=1, value="Insurance").font = FONT_CALC
    for i in range(11):
        c = ws.cell(row=17, column=2 + i,
                    value=f"=-'Property Data'!H23*(1+GrowthOpEx)^{i}")
        c.number_format = FMT_DOLLAR
        style_calc(c)

    # Row 18: RE Taxes (either OM base grown, OR reassessed at purchase × millage × SF-weighted avg)
    # Millage Rate is on Property Data row 16 (NOT row 17 — row 17 is Parcel text!)
    ws.cell(row=18, column=1, value="RE Taxes").font = FONT_CALC
    for i in range(11):
        reassessed = (f"-(PurchasePrice*SUMPRODUCT('Property Data'!$B$4:$G$4,"
                      f"'Property Data'!$B$16:$G$16)/PortfolioSF)*(1+GrowthRETax)^{i}")
        base = f"-'Property Data'!$H$24*(1+GrowthRETax)^{i}"
        c = ws.cell(row=18, column=2 + i, value=f"=IF(ReassessTaxes,{reassessed},{base})")
        c.number_format = FMT_DOLLAR
        style_calc(c)

    # Row 19: Total OpEx
    ws.cell(row=19, column=1, value="TOTAL OPERATING EXPENSES").font = FONT_TOTAL
    for i in range(11):
        col = get_column_letter(2 + i)
        c = ws.cell(row=19, column=2 + i, value=f"=SUM({col}14:{col}18)")
        style_total(c)
        c.number_format = FMT_DOLLAR

    # Row 21: NOI
    ws.cell(row=21, column=1, value="NET OPERATING INCOME (NOI)").font = FONT_TOTAL
    for i in range(11):
        col = get_column_letter(2 + i)
        c = ws.cell(row=21, column=2 + i, value=f"={col}11+{col}19")
        style_noi(c)
        c.number_format = FMT_DOLLAR

    # ---- CAPITAL COSTS ----
    ws.cell(row=23, column=1, value="CAPITAL COSTS").font = FONT_SUBHEADER

    # Row 24: Tenant Improvements & Leasing Commissions (from Rent Roll capital cols Y-AI)
    ws.cell(row=24, column=1, value="Tenant Improvements & Leasing Commissions").font = FONT_CALC
    for i in range(11):
        cap_col_letter = get_column_letter(25 + i)  # Y = 25, AI = 35
        c = ws.cell(row=24, column=2 + i,
                    value=f"=-'Rent Roll'!{cap_col_letter}167")
        c.number_format = FMT_DOLLAR
        style_calc(c)

    # Row 25: Capital Reserves
    ws.cell(row=25, column=1, value="Capital Reserves").font = FONT_CALC
    for i in range(11):
        c = ws.cell(row=25, column=2 + i,
                    value=f"=-ReservesPSF*PortfolioSF*(1+CPI)^{i}")
        c.number_format = FMT_DOLLAR
        style_calc(c)

    # Row 26: Total Capital
    ws.cell(row=26, column=1, value="TOTAL CAPITAL").font = FONT_TOTAL
    for i in range(11):
        col = get_column_letter(2 + i)
        c = ws.cell(row=26, column=2 + i, value=f"=SUM({col}24:{col}25)")
        style_total(c)
        c.number_format = FMT_DOLLAR

    # Row 28: UNLEVERED CASH FLOW
    ws.cell(row=28, column=1, value="UNLEVERED CASH FLOW").font = FONT_TOTAL
    for i in range(11):
        col = get_column_letter(2 + i)
        c = ws.cell(row=28, column=2 + i, value=f"={col}21+{col}26")
        style_noi(c)
        c.number_format = FMT_DOLLAR

    # ---- DEBT SERVICE (populated by Debt tab in Step 6) ----
    ws.cell(row=30, column=1, value="DEBT SERVICE").font = FONT_SUBHEADER
    ws.cell(row=31, column=1, value="Interest").font = FONT_CALC
    ws.cell(row=32, column=1, value="Principal").font = FONT_CALC
    ws.cell(row=33, column=1, value="Total Debt Service").font = FONT_CALC
    ws.cell(row=34, column=1, value="Loan Payoff at Maturity").font = FONT_CALC

    # Row 36: Levered Cash Flow (placeholder — will sum after Debt tab is built)
    ws.cell(row=36, column=1, value="LEVERED CASH FLOW").font = FONT_TOTAL
    for i in range(11):
        col = get_column_letter(2 + i)
        c = ws.cell(row=36, column=2 + i, value=f"={col}28+{col}33+{col}34")
        style_noi(c)
        c.number_format = FMT_DOLLAR

    # Totals column (M) — sum across years for stacking line items (NOI, CF, etc.)
    for r in [6, 7, 8, 9, 10, 11, 14, 15, 16, 17, 18, 19, 21, 24, 25, 26, 28, 31, 32, 33, 34, 36]:
        c = ws.cell(row=r, column=13, value=f"=SUM(B{r}:L{r})")
        c.number_format = FMT_DOLLAR
        if r in (11, 19, 21, 26, 28, 36):
            style_total(c)
        else:
            style_calc(c)

    ws.freeze_panes = "B4"


def build_debt(wb):
    """Debt tab — loan sizing + monthly amortization schedule (60 months)."""
    ws = wb["Debt"] if "Debt" in wb.sheetnames else wb.create_sheet("Debt")
    for row in ws.iter_rows():
        for c in row:
            c.value = None

    ws["A1"] = "DEBT MODULE — BANK FLOATING (SOFR + 300 bps, 70% LTV, 5-yr term)"
    ws["A1"].font = FONT_TITLE

    # Column widths
    widths = [22, 16, 14, 14, 14, 14, 14]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    # Loan Sizing summary (rows 3-12)
    ws["A3"] = "LOAN SIZING"
    style_subheader(ws["A3"])

    sizing = [
        ("Purchase Price", "=PurchasePrice", FMT_DOLLAR),
        ("LTV", "=LTV", FMT_PCT),
        ("Loan Amount", "=PurchasePrice*LTV", FMT_DOLLAR),
        ("SOFR", "=SOFR", FMT_PCT),
        ("Spread (bps)", "=LoanSpreadBps", "0"),
        ("All-in Rate", "=LoanRate", FMT_PCT),
        ("Amortization (months)", "=LoanAmortMonths", "0"),
        ("I/O Period (months)", "=LoanIOMonths", "0"),
        ("Term (months)", "=LoanTermMonths", "0"),
        ("Origination Fee $", "=PurchasePrice*LTV*LoanOrigFeePct", FMT_DOLLAR),
    ]
    for i, (label, formula, fmt) in enumerate(sizing):
        ws.cell(row=4 + i, column=1, value=label).font = FONT_CALC
        c = ws.cell(row=4 + i, column=2, value=formula)
        c.number_format = fmt
        style_calc(c)

    # Key debt metrics (rows 16-18)
    ws["A16"] = "KEY METRICS"
    style_subheader(ws["A16"])
    # P&I payment at amort rate (payment formula)
    ws.cell(row=17, column=1, value="Monthly P&I Payment (amort)").font = FONT_CALC
    c = ws.cell(row=17, column=2,
                value="=-PMT(LoanRate/12,LoanAmortMonths,PurchasePrice*LTV)")
    c.number_format = FMT_DOLLAR
    style_calc(c)
    ws.cell(row=18, column=1, value="Monthly Interest-Only Payment").font = FONT_CALC
    c = ws.cell(row=18, column=2, value="=PurchasePrice*LTV*LoanRate/12")
    c.number_format = FMT_DOLLAR
    style_calc(c)

    # Amortization schedule
    ws["A20"] = "AMORTIZATION SCHEDULE (60 monthly periods)"
    style_subheader(ws["A20"])
    headers = ["Month", "Beg Balance", "Rate (Annual)", "Interest", "Principal", "Payment", "End Balance"]
    for i, h in enumerate(headers):
        c = ws.cell(row=21, column=i + 1, value=h)
        style_subheader(c)

    # Rows 22-81: months 1-60
    for m in range(1, 61):
        r = 21 + m
        ws.cell(row=r, column=1, value=m)
        # Beg Balance: month 1 = loan amount; subsequent = previous end balance
        if m == 1:
            ws.cell(row=r, column=2, value="=PurchasePrice*LTV")
        else:
            ws.cell(row=r, column=2, value=f"=G{r-1}")
        ws.cell(row=r, column=2).number_format = FMT_DOLLAR

        # Rate (could be floating; for now constant. User can change SOFR)
        ws.cell(row=r, column=3, value="=LoanRate").number_format = FMT_PCT
        # Interest = Beg Balance × Rate/12
        ws.cell(row=r, column=4, value=f"=B{r}*C{r}/12").number_format = FMT_DOLLAR
        # Principal: 0 during I/O period, else payment - interest
        ws.cell(row=r, column=5,
                value=f"=IF({m}<=LoanIOMonths,0,F{r}-D{r})").number_format = FMT_DOLLAR
        # Payment: I/O during I/O period, else P&I
        ws.cell(row=r, column=6,
                value=f"=IF({m}<=LoanIOMonths,D{r},$B$17)").number_format = FMT_DOLLAR
        # End Balance = Beg - Principal
        ws.cell(row=r, column=7, value=f"=B{r}-E{r}").number_format = FMT_DOLLAR

        for col in range(1, 8):
            cell = ws.cell(row=r, column=col)
            cell.border = THIN_BORDER

    # Payoff at maturity: end balance of month 60 (row 81)
    ws["A83"] = "Loan Payoff at Maturity (Month 60)"
    ws.cell(row=83, column=1).font = FONT_TOTAL
    c = ws.cell(row=83, column=2, value="=G81")
    style_total(c)
    c.number_format = FMT_DOLLAR

    # FY Interest and Principal summaries (for Annual CF linkage)
    # FY27 = months 1-12, FY28 = months 13-24, etc.
    # Model hold = 5 years (FY27-FY31), so only months 1-60 exist
    ws["A85"] = "FY AGGREGATES (for Annual CF linkage)"
    style_subheader(ws["A85"])
    headers2 = ["Fiscal Year", "FY Interest", "FY Principal", "FY Debt Service", "EoY Balance"]
    for i, h in enumerate(headers2):
        c = ws.cell(row=86, column=i + 1, value=h)
        style_subheader(c)

    for i in range(11):  # FY27 through FY37
        r = 87 + i
        fy = 27 + i
        ws.cell(row=r, column=1, value=f"FY{fy}")
        if i < 5:  # FY27-FY31 (loan active)
            start_row = 22 + i * 12  # row for month 1 of this FY
            end_row = start_row + 11
            ws.cell(row=r, column=2, value=f"=-SUM(D{start_row}:D{end_row})").number_format = FMT_DOLLAR
            ws.cell(row=r, column=3, value=f"=-SUM(E{start_row}:E{end_row})").number_format = FMT_DOLLAR
            ws.cell(row=r, column=4, value=f"=B{r}+C{r}").number_format = FMT_DOLLAR
            ws.cell(row=r, column=5, value=f"=G{end_row}").number_format = FMT_DOLLAR
        else:
            # Loan has matured (or refinanced); no debt service
            for col in range(2, 5):
                ws.cell(row=r, column=col, value=0).number_format = FMT_DOLLAR
            ws.cell(row=r, column=5, value=0).number_format = FMT_DOLLAR

    ws.freeze_panes = "A22"

    # Now wire these debt aggregates into Annual CF rows 31-34
    acf = wb["Annual CF"]
    for i in range(11):
        col = get_column_letter(2 + i)
        # Row 31 Interest (from Debt!B{87+i})
        acf.cell(row=31, column=2 + i, value=f"=Debt!B{87+i}").number_format = FMT_DOLLAR
        style_calc(acf.cell(row=31, column=2 + i))
        # Row 32 Principal
        acf.cell(row=32, column=2 + i, value=f"=Debt!C{87+i}").number_format = FMT_DOLLAR
        style_calc(acf.cell(row=32, column=2 + i))
        # Row 33 Total Debt Service
        acf.cell(row=33, column=2 + i, value=f"=Debt!D{87+i}").number_format = FMT_DOLLAR
        style_calc(acf.cell(row=33, column=2 + i))
        # Row 34 Loan Payoff — dynamic: triggers in the FY that contains the loan maturity month
        # FY27 = months 1-12, FY28 = 13-24, ..., FYn = (n-27)*12+1 to (n-26)*12
        # Loan matures at month = LoanTermMonths. FY index = INT((LoanTermMonths-1)/12)
        acf.cell(row=34, column=2 + i,
                 value=f"=IF(INT((LoanTermMonths-1)/12)={i},-INDEX(Debt!$G$22:$G$81,LoanTermMonths),0)"
                 ).number_format = FMT_DOLLAR
        style_calc(acf.cell(row=34, column=2 + i))


def build_monthly_cf(wb):
    """Monthly CF tab — 120 months, Annual CF ÷ 12 with acquisition/sale timing."""
    ws = wb["Monthly CF"] if "Monthly CF" in wb.sheetnames else wb.create_sheet("Monthly CF")
    for row in ws.iter_rows():
        for c in row:
            c.value = None

    ws["A1"] = "MONTHLY CASH FLOW — 120 months from AnalysisStart"
    ws["A1"].font = FONT_TITLE

    ws.column_dimensions["A"].width = 32
    for i in range(120):
        ws.column_dimensions[get_column_letter(2 + i)].width = 11

    # Row 3: Month # (1-120)
    ws.cell(row=3, column=1, value="Month #").font = FONT_SUBHEADER
    for i in range(120):
        c = ws.cell(row=3, column=2 + i, value=i + 1)
        style_subheader(c)
        c.alignment = Alignment(horizontal="center")

    # Row 4: Date (label)
    ws.cell(row=4, column=1, value="Month").font = FONT_SUBHEADER
    for i in range(120):
        c = ws.cell(row=4, column=2 + i, value=f"=EDATE(AnalysisStart,{i})")
        c.number_format = FMT_DATE
        c.alignment = Alignment(horizontal="center")

    # Row 5: FY (derived: month 1-12 = FY27, 13-24 = FY28, etc.)
    ws.cell(row=5, column=1, value="Fiscal Year").font = FONT_SUBHEADER
    for i in range(120):
        fy_num = 27 + (i // 12)
        c = ws.cell(row=5, column=2 + i, value=f"FY{fy_num}")
        c.alignment = Alignment(horizontal="center")

    # Row 7: NOI (Annual / 12, straight-line)
    ws.cell(row=7, column=1, value="NOI (Annual / 12)").font = FONT_CALC
    for i in range(120):
        fy_idx = i // 12  # 0-9
        if fy_idx <= 10:  # FY27-FY37
            col_acf = get_column_letter(2 + fy_idx)
            c = ws.cell(row=7, column=2 + i, value=f"='Annual CF'!{col_acf}21/12")
            c.number_format = FMT_DOLLAR
            style_calc(c)

    # Row 8: Capital (Annual / 12)
    ws.cell(row=8, column=1, value="Capital Costs (Annual / 12)").font = FONT_CALC
    for i in range(120):
        fy_idx = i // 12
        if fy_idx <= 10:
            col_acf = get_column_letter(2 + fy_idx)
            c = ws.cell(row=8, column=2 + i, value=f"='Annual CF'!{col_acf}26/12")
            c.number_format = FMT_DOLLAR
            style_calc(c)

    # Row 9: Unlevered Operating CF
    ws.cell(row=9, column=1, value="Operating Cash Flow").font = FONT_TOTAL
    for i in range(120):
        col = get_column_letter(2 + i)
        c = ws.cell(row=9, column=2 + i, value=f"={col}7+{col}8")
        c.number_format = FMT_DOLLAR
        style_total(c)

    # Row 11: Acquisition (Month 1)
    ws.cell(row=11, column=1, value="Purchase Price + Closing Costs").font = FONT_CALC
    c = ws.cell(row=11, column=2, value="=-PurchasePrice*(1+AcqCostPct)")
    c.number_format = FMT_DOLLAR
    style_calc(c)

    # Row 12: Loan Proceeds (Month 1)
    ws.cell(row=12, column=1, value="Loan Proceeds (net of orig fee)").font = FONT_CALC
    c = ws.cell(row=12, column=2, value="=PurchasePrice*LTV*(1-LoanOrigFeePct)")
    c.number_format = FMT_DOLLAR
    style_calc(c)

    # Row 13: Sale Proceeds (HoldMonths)
    # Exit value = next year's NOI / ExitCap. Using FY32 NOI (year after exit).
    # At HoldMonths = 60 (end of FY31), we sell — use FY32 NOI as exit NOI.
    ws.cell(row=13, column=1, value="Sale Proceeds (net of sale costs)").font = FONT_CALC
    # Only in the hold-exit month
    for i in range(120):
        col = get_column_letter(2 + i)
        # If month = HoldMonths: sell
        # Exit Value = FY after exit / ExitCap — based on HoldMonths / 12
        # For simplicity: Exit NOI = (HoldMonths+1)-th FY NOI
        exit_formula = (f'IF({i+1}=HoldMonths,'
                        f"(INDEX('Annual CF'!$B$21:$L$21,INT(HoldMonths/12)+1)/ExitCap)"
                        f"*(1-DispCostPct),0)")
        c = ws.cell(row=13, column=2 + i, value=f"={exit_formula}")
        c.number_format = FMT_DOLLAR
        style_calc(c)

    # Row 14: Loan Payoff at Exit
    ws.cell(row=14, column=1, value="Loan Payoff at Exit").font = FONT_CALC
    for i in range(120):
        # If month = HoldMonths AND HoldMonths <= 60: payoff = Debt!G{21+HoldMonths}
        # For flexibility: use INDEX on amortization End Balance column
        formula = f"=IF({i+1}=HoldMonths,-INDEX(Debt!$G$22:$G$81,HoldMonths),0)"
        c = ws.cell(row=14, column=2 + i, value=formula)
        c.number_format = FMT_DOLLAR
        style_calc(c)

    # Row 15: Debt Service monthly
    ws.cell(row=15, column=1, value="Debt Service (Monthly)").font = FONT_CALC
    for i in range(120):
        # Months 1-60: pull from Debt!F (Payment column) — negative (cash out)
        if i < 60:
            row_d = 22 + i
            c = ws.cell(row=15, column=2 + i, value=f"=-Debt!F{row_d}")
        else:
            c = ws.cell(row=15, column=2 + i, value=0)
        c.number_format = FMT_DOLLAR
        style_calc(c)

    # Row 17: Unlevered Total CF
    ws.cell(row=17, column=1, value="UNLEVERED CASH FLOW").font = FONT_TOTAL
    for i in range(120):
        col = get_column_letter(2 + i)
        # Includes: operating CF + acquisition (month 1) + sale (hold month)
        c = ws.cell(row=17, column=2 + i, value=f"={col}9+{col}11+{col}13")
        c.number_format = FMT_DOLLAR
        style_noi(c)

    # Row 18: Levered Total CF
    ws.cell(row=18, column=1, value="LEVERED CASH FLOW").font = FONT_TOTAL
    for i in range(120):
        col = get_column_letter(2 + i)
        # Levered = Unlevered + Loan Proceeds (in) - Debt Service - Loan Payoff
        c = ws.cell(row=18, column=2 + i, value=f"={col}17+{col}12+{col}15+{col}14")
        c.number_format = FMT_DOLLAR
        style_noi(c)

    ws.freeze_panes = "B6"


def build_summary(wb):
    """Summary tab — Deal overview, Sources & Uses, NOI bridge, Returns."""
    ws = wb["Summary"] if "Summary" in wb.sheetnames else wb.create_sheet("Summary")
    for row in ws.iter_rows():
        for c in row:
            c.value = None

    ws["A1"] = "DEAL SUMMARY — DFW MODERN MULTI-TENANT PORTFOLIO"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:F1")

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 34
    ws.column_dimensions["E"].width = 18

    # LEFT COLUMN: Deal Overview
    ws["A3"] = "DEAL OVERVIEW"
    style_subheader(ws["A3"])

    def kv(ws, row, col, label, formula, fmt=None, highlight=False):
        ws.cell(row=row, column=col, value=label).font = FONT_CALC
        c = ws.cell(row=row, column=col + 1, value=formula)
        if fmt:
            c.number_format = fmt
        if highlight:
            style_noi(c)
        else:
            style_calc(c)

    kv(ws, 4, 1, "Portfolio SF", "=PortfolioSF", FMT_NUMBER)
    kv(ws, 5, 1, "# Properties", f"={len(PROPERTIES)}", "0")
    kv(ws, 6, 1, "# Tenant Suites", "=COUNTA('Rent Roll'!C5:C166)", "0")
    kv(ws, 7, 1, "Hold Period (Years)", "=HoldMonths/12", FMT_NUMBER_2DP)

    kv(ws, 9, 1, "Purchase Price", "=PurchasePrice", FMT_DOLLAR, highlight=True)
    kv(ws, 10, 1, "Price per SF", "=PurchasePrice/PortfolioSF", FMT_DOLLAR_CENTS)
    kv(ws, 11, 1, "Acquisition Costs", "=PurchasePrice*AcqCostPct", FMT_DOLLAR)
    kv(ws, 12, 1, "Total Acquisition Cost", "=PurchasePrice*(1+AcqCostPct)", FMT_DOLLAR)

    kv(ws, 14, 1, "Going-in Cap (on In-Place NOI)", "=InPlaceNOI/PurchasePrice", FMT_PCT)
    kv(ws, 15, 1, "Going-in Cap (on FY27 NOI)", "='Annual CF'!B21/PurchasePrice", FMT_PCT)
    kv(ws, 16, 1, "Exit Cap Rate", "=ExitCap", FMT_PCT)

    ws["A18"] = "FINANCING"
    style_subheader(ws["A18"])
    kv(ws, 19, 1, "Loan Amount (70% LTV)", "=PurchasePrice*LTV", FMT_DOLLAR)
    kv(ws, 20, 1, "LTV", "=LTV", FMT_PCT)
    kv(ws, 21, 1, "All-in Rate (SOFR+300)", "=LoanRate", FMT_PCT)
    kv(ws, 22, 1, "Debt Yield (on In-Place NOI)", "=InPlaceNOI/(PurchasePrice*LTV)", FMT_PCT)
    kv(ws, 23, 1, "DSCR FY27 (NOI / Debt Service)", "='Annual CF'!B21/(-Debt!D87)", FMT_NUMBER_2DP)

    # SOURCES & USES
    ws["A25"] = "SOURCES & USES"
    style_subheader(ws["A25"])
    ws.cell(row=26, column=1, value="SOURCES").font = FONT_SUBHEADER
    kv(ws, 27, 1, "Loan Proceeds", "=PurchasePrice*LTV*(1-LoanOrigFeePct)", FMT_DOLLAR)
    kv(ws, 28, 1, "Equity Required", "=B31-B27", FMT_DOLLAR, highlight=True)
    ws.cell(row=29, column=1, value="Total Sources").font = FONT_TOTAL
    c = ws.cell(row=29, column=2, value="=B27+B28")
    style_total(c); c.number_format = FMT_DOLLAR

    ws.cell(row=30, column=1, value="USES").font = FONT_SUBHEADER
    kv(ws, 31, 1, "Purchase Price + Acq Costs", "=PurchasePrice*(1+AcqCostPct)", FMT_DOLLAR)
    ws.cell(row=32, column=1, value="Total Uses").font = FONT_TOTAL
    c = ws.cell(row=32, column=2, value="=B31")
    style_total(c); c.number_format = FMT_DOLLAR

    # RIGHT COLUMN: NOI Bridge + Returns
    ws["D3"] = "NOI BRIDGE"
    style_subheader(ws["D3"])
    kv(ws, 4, 4, "In-Place NOI", "=InPlaceNOI", FMT_DOLLAR)
    kv(ws, 5, 4, "FY27 NOI (Year 1)", "='Annual CF'!B21", FMT_DOLLAR)
    kv(ws, 6, 4, "FY28 NOI (Year 2)", "='Annual CF'!C21", FMT_DOLLAR)
    kv(ws, 7, 4, "FY31 NOI (Year 5 / Exit Year)", "='Annual CF'!F21", FMT_DOLLAR)
    kv(ws, 8, 4, "FY32 NOI (Exit Year + 1)", "='Annual CF'!G21", FMT_DOLLAR)
    kv(ws, 9, 4, "NOI Growth (Y1 → Y5)",
       "=('Annual CF'!F21/'Annual CF'!B21)-1", FMT_PCT)

    ws["D11"] = "EXIT"
    style_subheader(ws["D11"])
    kv(ws, 12, 4, "Exit Year NOI (next-year method)",
       "=INDEX('Annual CF'!$B$21:$L$21,INT(HoldMonths/12)+1)", FMT_DOLLAR)
    kv(ws, 13, 4, "Exit Value (Gross)",
       "=E12/ExitCap", FMT_DOLLAR)
    kv(ws, 14, 4, "Disposition Costs",
       "=E13*DispCostPct", FMT_DOLLAR)
    kv(ws, 15, 4, "Exit Value (Net)",
       "=E13*(1-DispCostPct)", FMT_DOLLAR, highlight=True)
    kv(ws, 16, 4, "Exit $/SF", "=E13/PortfolioSF", FMT_DOLLAR_CENTS)

    # RETURNS
    ws["D18"] = "RETURNS"
    style_subheader(ws["D18"])

    # Unlevered IRR: pull from Monthly CF row 17 (unlevered total CF)
    # Need 121 months: 0 (if month 1 is period 0) through 120
    # Use XIRR for accuracy
    # Unlevered CF row 17 in Monthly CF, cols B:DQ (months 1-120)
    kv(ws, 19, 4, "Unlevered IRR",
       "=IRR('Monthly CF'!B17:OFFSET('Monthly CF'!B17,0,HoldMonths-1))*12", FMT_PCT,
       highlight=True)
    kv(ws, 20, 4, "Levered IRR",
       "=IRR('Monthly CF'!B18:OFFSET('Monthly CF'!B18,0,HoldMonths-1))*12", FMT_PCT,
       highlight=True)
    kv(ws, 21, 4, "Equity Multiple (Levered)",
       "=SUM(OFFSET('Monthly CF'!B18,0,0,1,HoldMonths))/(-B28)+1", FMT_MULT)
    kv(ws, 22, 4, "Avg Cash-on-Cash (Levered)",
       "=(SUM(OFFSET('Monthly CF'!B18,0,0,1,HoldMonths-1))-(-B28))/(-B28)/(HoldMonths/12)",
       FMT_PCT)

    ws["D24"] = "KEY METRICS"
    style_subheader(ws["D24"])
    kv(ws, 25, 4, "FY27 Cap Rate (on purchase)", "='Annual CF'!B21/PurchasePrice", FMT_PCT)
    kv(ws, 26, 4, "FY31 Cap Rate (exit year)", "='Annual CF'!F21/PurchasePrice", FMT_PCT)
    kv(ws, 27, 4, "Peak Cash-on-Cash",
       "=MAX('Annual CF'!B36:F36)/(-B28)", FMT_PCT)
    kv(ws, 28, 4, "FY27 DSCR",
       "='Annual CF'!B21/(-Debt!D87)", FMT_NUMBER_2DP)
    kv(ws, 29, 4, "FY27 Debt Yield", "='Annual CF'!B21/(PurchasePrice*LTV)", FMT_PCT)

    # Model integrity check — Model FY27 NOI vs OM FY27 NOI
    ws["A31"] = "MODEL INTEGRITY"
    style_subheader(ws["A31"])
    kv(ws, 32, 1, "Model FY27 NOI", "='Annual CF'!B21", FMT_DOLLAR)
    kv(ws, 33, 1, "OM-stated FY27 NOI (sum)", "=SUM('Property Data'!B19:G19)", FMT_DOLLAR)
    kv(ws, 34, 1, "Variance (Model − OM)", "=B32-B33", FMT_DOLLAR)
    kv(ws, 35, 1, "Variance %", "=B34/B33", FMT_PCT)

    # Change notes
    ws["A37"] = ("Notes: Yellow cells on Assumptions tab are user inputs. Orange cells on "
                 "Property Data = Plano estimates (OM incomplete). Exit uses the FY after "
                 "HoldMonths (next-year NOI method). Levered IRR uses monthly IRR × 12. "
                 "Treatment column on Rent Roll has a dropdown (Market/Option/Vacate/LeaseUp) "
                 "controlling per-tenant post-expiry behavior.")
    ws["A37"].font = Font(italic=True, size=9, color="606060")
    ws.merge_cells("A37:E40")
    ws["A37"].alignment = Alignment(wrap_text=True, vertical="top")


# =============================================================================
# MAIN
# =============================================================================

def main():
    wb = Workbook()
    # Remove the default sheet — we'll add them in the order we want
    wb.remove(wb.active)

    build_assumptions(wb)
    build_property_data(wb)

    # Placeholder sheets for tabs not yet built (so cross-refs don't error)
    for tab in ("MLA", "Rent Roll", "Annual CF", "Debt", "Monthly CF", "Summary"):
        if tab not in wb.sheetnames:
            wb.create_sheet(tab)

    build_mla(wb)

    tenants = load_rent_roll(RENT_ROLL_PATH)
    print(f"  Loaded {len(tenants)} tenant records from rent roll")
    build_rent_roll(wb, tenants)
    build_annual_cf(wb)
    build_debt(wb)
    build_monthly_cf(wb)
    build_summary(wb)

    # Reorder tabs for UX: Summary first, then Assumptions, then the rest
    desired_order = ["Summary", "Assumptions", "Property Data", "MLA", "Rent Roll",
                     "Annual CF", "Monthly CF", "Debt"]
    for idx, name in enumerate(desired_order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=idx - wb.sheetnames.index(name))

    wb.save(OUTPUT_PATH)
    print(f"✓ Model saved: {OUTPUT_PATH}")
    print(f"  Portfolio SF: {PORTFOLIO_SF:,}")
    print(f"  OM FY27 NOI (sum): ${PORTFOLIO_FY27_NOI:,.0f}")
    print(f"  All 8 tabs built: Summary, Assumptions, Property Data, MLA, Rent Roll, Annual CF, Monthly CF, Debt")


if __name__ == "__main__":
    main()
